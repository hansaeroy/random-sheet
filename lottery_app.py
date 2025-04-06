import streamlit as st
import pandas as pd
import secrets
import random
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# 제비뽑기 함수 정의
def create_random_seating_assignment(uploaded_file):
    try:
        # 엑셀 파일 읽기
        names_df = pd.read_excel(uploaded_file)
        
        # 명단 시트에서 이름 목록 추출 (모든 열에서)
        names = []
        for column in names_df.columns:
            names.extend(names_df[column].dropna().tolist())
        
        names = [str(name) for name in names if isinstance(name, str) or isinstance(name, (int, float))]
        
        if len(names) == 0:
            st.error("명단에서 이름을 찾을 수 없습니다.")
            return None
        
        # 좌석 번호 생성
        regular_seats = list(range(1, 222))  # 일반 좌석 1-221
        chair_seats = [f"의자{i}" for i in range(1, 42)]  # 의자1-의자41
        
        # 좌석 수와 명단 수 확인
        if len(names) > len(regular_seats) + len(chair_seats):
            st.error(f"명단({len(names)}명)이 좌석 수({len(regular_seats) + len(chair_seats)}개)보다 많습니다.")
            return None
        
        # 데이터프레임 생성을 위한 데이터 준비
        result_data = []
        
        # 암호학적으로 안전한 난수 생성기를 사용하여 각 이름에 랜덤 값 할당
        for name in names:
            random_value = secrets.randbelow(1000000) / 1000000
            result_data.append({
                '이름': name,
                '랜덤값': random_value
            })
        
        # 결과 데이터프레임 생성 및 정렬
        result_df = pd.DataFrame(result_data)
        result_df = result_df.sort_values(by='랜덤값')
        
        # 일반 좌석만 섞기
        random.shuffle(regular_seats)
        
        # 좌석 번호 할당 (랜덤 순서대로, 일반 좌석 먼저 배정)
        needed_regular_seats = min(len(names), len(regular_seats))
        needed_chair_seats = max(0, len(names) - needed_regular_seats)
        
        # 일반 좌석과 필요한 경우 의자 좌석 할당
        assigned_seats = regular_seats[:needed_regular_seats]
        if needed_chair_seats > 0:
            assigned_seats.extend(chair_seats[:needed_chair_seats])
        
        # 결과 데이터프레임에 당첨번호 할당
        result_df['당첨번호'] = assigned_seats[:len(names)]
        
        # 이름 기준으로 다시 정렬 (가나다순)
        result_df_sorted = result_df.sort_values(by='이름').reset_index(drop=True)
        
        return {
            'result_df': result_df_sorted,
            'names': sorted(names),
            'needed_regular_seats': needed_regular_seats,
            'needed_chair_seats': needed_chair_seats
        }
        
    except Exception as e:
        st.error(f"오류 발생: {e}")
        return None

# 결과 엑셀 파일 생성 함수
def create_result_excel(results):
    # 결과 데이터프레임
    df = results['result_df']
    
    # 새 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "제비뽑기 결과"
    
    # 페이지 설정
    ws.page_setup.paperSize = 9  # A4 용지
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.horizontalCentered = True
    ws.print_options.horizontalCentered = True
    
    # 여백 설정
    ws.page_margins = PageMargins(bottom=0.4)
    
    # 맞춤 설정
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # 스타일 정의
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    light_blue_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    
    # 현재 날짜
    today = datetime.now().strftime('%m월 %d일')
    
    # 섹션별 행 수와 열 수
    rows_per_section = 30
    cols_per_section = 3
    
    total_persons = len(df)
    persons_per_section = rows_per_section * cols_per_section
    num_sections = (total_persons + persons_per_section - 1) // persons_per_section
    
    # 현재 행 위치
    current_row = 1
    
    # 섹션별로 데이터 추가
    for section_idx in range(num_sections):
        section_start_row = current_row
        
        # 제목 행
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        title_cell = ws.cell(row=current_row, column=1, value=f"제비뽑기 당첨 결과 {section_idx+1}")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 32
        current_row += 1
        
        # 날짜 행
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        date_cell = ws.cell(row=current_row, column=1, value=f"날짜: {today}")
        date_cell.font = Font(bold=True)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 24
        
        # (가나다순) 텍스트
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        sort_cell = ws.cell(row=current_row, column=5, value="(가나다순)")
        sort_cell.font = Font(bold=True)
        sort_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # 헤더 행
        headers = ["이 름", "당첨번호", "이 름", "당첨번호", "이 름", "당첨번호"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=current_row, column=i+1, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
        # 해당 섹션의 데이터 범위
        start_idx = section_idx * persons_per_section
        end_idx = min(start_idx + persons_per_section, total_persons)
        section_data = df.iloc[start_idx:end_idx].reset_index(drop=True)
        
        # 최대 행 인덱스 추적
        max_row_idx = -1
        
        # 섹션 데이터 추가
        for idx, row in section_data.iterrows():
            col_set = idx // rows_per_section
            row_idx = idx % rows_per_section
            max_row_idx = max(max_row_idx, row_idx)
            
            # 열 인덱스 계산
            col_idx = col_set * 2 + 1
            
            # 현재 데이터 행 위치
            data_row = current_row + row_idx
            
            # 이름 열과 당첨번호 열
            name_cell = ws.cell(row=data_row, column=col_idx)
            num_cell = ws.cell(row=data_row, column=col_idx + 1)
            
            # 스타일 설정
            name_cell.border = thin_border
            num_cell.border = thin_border
            num_cell.fill = light_blue_fill
            name_cell.font = Font(bold=True)
            num_cell.font = Font(bold=True)
            name_cell.alignment = Alignment(horizontal='center', vertical='center')
            num_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 데이터 설정
            name_cell.value = row['이름']
            num_cell.value = row['당첨번호']
            
            ws.row_dimensions[data_row].height = 22.80
        
        # 빈 데이터 처리
        if max_row_idx == -1:
            max_row_idx = 0
            
        # 섹션 마지막 행 계산
        section_end_row = current_row + max_row_idx
        
        # 섹션 테두리 추가
        for r in range(section_start_row, section_end_row + 1):
            for c in range(1, 7):
                if r == section_start_row or r == section_end_row or c == 1 or c == 6:
                    cell = ws.cell(row=r, column=c)
                    if cell.border:
                        # 테두리 처리 로직
                        if (r == section_start_row and c == 1):  # 좌상단 모서리
                            cell.border = Border(
                                left=Side(style='medium'), 
                                right=cell.border.right, 
                                top=Side(style='medium'), 
                                bottom=cell.border.bottom
                            )
                        elif (r == section_start_row and c == 6):  # 우상단 모서리
                            cell.border = Border(
                                left=cell.border.left, 
                                right=Side(style='medium'), 
                                top=Side(style='medium'), 
                                bottom=cell.border.bottom
                            )
                        elif (r == section_end_row and c == 1):  # 좌하단 모서리
                            cell.border = Border(
                                left=Side(style='medium'), 
                                right=cell.border.right, 
                                top=cell.border.top, 
                                bottom=Side(style='medium')
                            )
                        elif (r == section_end_row and c == 6):  # 우하단 모서리
                            cell.border = Border(
                                left=cell.border.left, 
                                right=Side(style='medium'), 
                                top=cell.border.top, 
                                bottom=Side(style='medium')
                            )
                        elif r == section_start_row:  # 상단 테두리
                            cell.border = Border(
                                left=cell.border.left, 
                                right=cell.border.right, 
                                top=Side(style='medium'), 
                                bottom=cell.border.bottom
                            )
                        elif r == section_end_row:  # 하단 테두리
                            cell.border = Border(
                                left=cell.border.left, 
                                right=cell.border.right, 
                                top=cell.border.top, 
                                bottom=Side(style='medium')
                            )
                        elif c == 1:  # 좌측 테두리
                            cell.border = Border(
                                left=Side(style='medium'), 
                                right=cell.border.right, 
                                top=cell.border.top, 
                                bottom=cell.border.bottom
                            )
                        elif c == 6:  # 우측 테두리
                            cell.border = Border(
                                left=cell.border.left, 
                                right=Side(style='medium'), 
                                top=cell.border.top, 
                                bottom=cell.border.bottom
                            )
                    else:
                        cell.border = medium_border
        
        # 다음 섹션 위치 업데이트
        current_row = section_end_row + 1
    
    # 열 너비 조정
    for i in range(1, 7):
        col_letter = get_column_letter(i)
        if i % 2 == 1:  # 홀수 열 (이름)
            ws.column_dimensions[col_letter].width = 15
        else:  # 짝수 열 (당첨번호)
            ws.column_dimensions[col_letter].width = 12
    
    # 엑셀 파일을 바이트로 변환
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# 페이지 설정
st.set_page_config(page_title="제비뽑기 프로그램", page_icon="🎯", layout="wide")

# 간결화된 CSS 스타일
st.markdown("""
<style>
body { color: rgba(250, 250, 250, 0.9) !important; }
p, ol, ul, label, div { color: rgba(250, 250, 250, 0.9) !important; }
h1 { color: #ffffff !important; text-align: center; margin-bottom: 2rem; }
h2, h3, h4 { color: #ffffff !important; }

.stButton > button {
    background-color: #4CAF50 !important;
    color: white !important;
    font-size: 20px !important;
    padding: 12px 24px !important;
    border-radius: 8px !important;
    border: none !important;
    cursor: pointer !important;
    margin: 20px 0 !important;
    display: block !important;
    width: 100% !important;
    transition: all 0.3s !important;
}
.stButton > button:hover { 
    background-color: #3e8e41 !important; 
    box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important; 
    transform: translateY(-2px) !important; 
}

[data-testid="stDownloadButton"] > button {
    background-color: #4CAF50 !important;
    color: white !important;
    font-size: 20px !important;
    padding: 12px 24px !important;
    border-radius: 8px !important;
    border: none !important;
    cursor: pointer !important;
    margin: 20px 0 !important;
    display: block !important;
    width: 100% !important;
    transition: all 0.3s !important;
    height: auto !important;
    line-height: 1.6 !important;
}
[data-testid="stDownloadButton"] > button:hover { 
    background-color: #3e8e41 !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important;
    transform: translateY(-2px) !important; 
}
[data-testid="stDownloadButton"] { 
    margin-top: 20px !important; margin-bottom: 20px !important;
    display: block !important; width: 100% !important; 
}

.css-1cpxqw2, [data-testid="stFileUploader"] {
    border: 2px dashed #4CAF50 !important;
    border-radius: 10px !important;
    padding: 30px !important;
    text-align: center !important;
    transition: all 0.3s !important;
    background-color: rgba(255, 255, 255, 0.05) !important;
    min-height: 280px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: center !important;
    align-items: center !important;
}
.css-1cpxqw2:hover, [data-testid="stFileUploader"]:hover {
    border-color: #ffffff !important;
    background-color: rgba(255, 255, 255, 0.1) !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.3) !important;
}

.download-container {
    padding: 20px !important;
    background-color: rgba(30, 30, 30, 0.6) !important;
    border-radius: 10px !important;
    margin: 0 !important;
    text-align: center !important;
    border: 1px solid rgba(255, 255, 255, 0.2) !important;
    min-height: 280px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: center !important;
    align-items: center !important;
    width: 100% !important;
}
.equal-height-container {
    height: 100% !important;
    display: flex !important;
    flex-direction: column !important;
}

.info-box {
    background-color: rgba(30, 30, 30, 0.6) !important;
    padding: 15px !important;
    border-radius: 8px !important;
    margin: 20px 0 !important;
    border-left: 5px solid #4CAF50 !important;
    color: rgba(250, 250, 250, 0.9) !important;
}
.element-container.st-success {
    background-color: rgba(76, 175, 80, 0.1) !important;
    padding: 15px !important;
    border-radius: 8px !important;
    border-left: 5px solid #4CAF50 !important;
    margin: 20px 0 !important;
}
.result-summary {
    margin-top: 20px !important;
    background-color: rgba(30, 30, 30, 0.5) !important;
    padding: 15px !important;
    border-radius: 8px !important;
    border-left: 5px solid #4CAF50 !important;
}
</style>
""", unsafe_allow_html=True)

# 제목과 설명
st.title("🎯 신학교 제비뽑기 프로그램")
st.markdown("""
<div class="info-box">
    <h3>사용 방법</h3>
    <ol>
        <li>명단이 있는 엑셀 파일(.xlsx, .xls)을 업로드하세요.</li>
        <li>'제비뽑기 실행' 버튼을 클릭하세요.</li>
        <li>결과가 생성되면 '결과 파일 다운로드' 버튼을 클릭하여 저장하세요.</li>
        <li>Excel 파일이 한 페이지에 모든 결과가 나타나도록 설정되었습니다.</li>
    </ol>
</div>
""", unsafe_allow_html=True)

# 화면 2단 분할
col1, col2 = st.columns([1, 1])

with col1:
    st.markdown('<div class="equal-height-container">', unsafe_allow_html=True)
    
    # 파일 업로드 위젯
    uploaded_file = st.file_uploader("명단이 있는 엑셀 파일을 업로드하세요", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        st.success(f"파일 '{uploaded_file.name}'이 업로드되었습니다.")
    
    # 제비뽑기 실행 버튼
    if st.button("제비뽑기 실행"):
        with st.spinner("제비뽑기 진행 중..."):
            results = create_random_seating_assignment(uploaded_file)
            
            if results:
                st.session_state.results = results
                st.session_state.excel_data = create_result_excel(results)
                st.session_state.execution_completed = True
                
                # 결과 요약
                needed_regular = results['needed_regular_seats']
                needed_chair = results['needed_chair_seats']
                total_people = needed_regular + needed_chair
                
                st.success(f"✅ 제비뽑기 완료! 총 {total_people}명 배정 ({needed_regular}개 일반 좌석, {needed_chair}개 의자 좌석)")
    
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="equal-height-container" style="width:100%;">', unsafe_allow_html=True)
    
    if 'execution_completed' in st.session_state and st.session_state.execution_completed:
        st.markdown("""
        <div class="download-container">
            <h3>결과 다운로드</h3>
            <p>제비뽑기가 완료되었습니다!</p>
        """, unsafe_allow_html=True)
        
        # 파일명 생성
        file_name = f"제비뽑기_결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Excel 파일 다운로드 버튼
        st.download_button(
            label="📥 결과 파일 다운로드",
            data=st.session_state.excel_data,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel",
            help="결과 Excel 파일을 다운로드합니다.",
            use_container_width=True
        )
        
        st.write(f"일반 좌석: {st.session_state.results['needed_regular_seats']}개")
        st.write(f"의자 좌석: {st.session_state.results['needed_chair_seats']}개")
        
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="download-container">
            <h3>결과 다운로드</h3>
            <p>왼쪽에서 파일을 업로드하고 제비뽑기를 실행하면 여기에 다운로드 버튼이 나타납니다.</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)

# 프로그램 설명
with st.expander("제비뽑기 프로그램 상세 설명"):
    st.markdown("""
    ### 제비뽑기 프로그램 특징
    
    **좌석 배정 방식:**
    - 일반 좌석(1~221)이 랜덤하게 배정됩니다.
    - 인원이 221명을 초과하는 경우에만 의자 좌석이 배정됩니다.
    - 의자 좌석은 의자1부터 순차적으로 필요한 만큼만 배정됩니다.
    
    **결과 파일 형식:**
    - Excel 파일로 다운로드됩니다.
    - 모든 결과가 하나의 시트에 표시됩니다.
    - 각 섹션별로 제목과 헤더가 추가되어 구분이 용이합니다.
    - 세로 방향 인쇄로 설정되어 있으며, 페이지 여백이 가로 가운데 맞춤으로 조정되었습니다.
    - 모든 텍스트는 굵게 처리되고 중앙 정렬됩니다.
    - 당첨번호 열은 연한 파란색 배경으로 표시됩니다.
    """)

# 푸터
st.markdown("---")
st.markdown("<p style='text-align: center; color: rgba(250, 250, 250, 0.7);'>© 2025 신학교 제비뽑기 프로그램 | 제작자: 여치형</p>", unsafe_allow_html=True)
