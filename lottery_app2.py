import streamlit as st
import pandas as pd
import secrets
import random
import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import openpyxl
from openpyxl.utils.cell import range_boundaries
from copy import copy
from openpyxl.drawing.image import Image

# 앱 디렉토리에 좌석 배치표 파일 저장
SEATING_CHART_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "seating_chart.xlsx")

# 픽셀을 Excel 열 너비 단위로 정확하게 변환하는 함수
def pixels_to_excel_width(pixels):
    # 공식: Excel 열 너비 = (픽셀 - 셀 패딩) / 문자 폭 계수
    padding = 5
    char_width = 9.5  # 81픽셀에서 61픽셀로 줄이기 위해 조정된 값
    
    return (pixels - padding) / char_width

def create_random_seating_assignment(uploaded_file):
    try:
        # 엑셀 파일 읽기
        names_df = pd.read_excel(uploaded_file)
        
        # 이름과 그룹 정보를 추출
        persons = []
        
        # 기관 정보 추적
        current_group = None
        
        # 명단 리스트 - 누락 검사용
        all_names = []
        
        # 모든 행 순회
        for i in range(len(names_df)):
            # 첫 번째 열 체크 - 기관명 확인
            group_cell = names_df.iloc[i, 0]
            if pd.notna(group_cell) and isinstance(group_cell, str):
                group_str = str(group_cell).strip()
                # 새로운 기관 발견 (남, 여, 청, 안나, 디모데, 사모회 등)
                if any(group_str.endswith(marker) for marker in ['남', '여', '청', '안나']) or group_str in ['디모데', '사모회']:
                    current_group = group_str
            
            # 이름 열 순회 (첫번째 열 제외, 마지막 열(합계) 제외)
            for j in range(1, len(names_df.columns) - 1):
                # 셀 값 가져오기
                value = names_df.iloc[i, j]
                
                # 값이 존재하는 경우만 처리
                if pd.notna(value):
                    value_str = str(value).strip()
                    
                    # 이름인지 확인 (공백 제거 후 길이 2~6자, 숫자나 특정 키워드 아님)
                    is_name = (
                        2 <= len(value_str) <= 6 and 
                        not value_str.isdigit() and 
                        not "." in value_str and  # 소수점 있는 숫자 제외
                        value_str not in ["기관", "합계", "명단", "NaT"] and
                        not (value_str in ["남", "여", "청", "안나", "디모데", "사모회"])
                    )
                    
                    if is_name and current_group:
                        # 이름 목록에 추가
                        all_names.append(value_str)
                        
                        # 결과 목록에 추가
                        persons.append({
                            '이름': value_str,
                            '그룹': current_group
                        })
        
        # 중복 제거 (동명이인은 유지 - 그룹과 함께 고려)
        unique_persons = []
        seen = set()
        
        for person in persons:
            # 이름과 그룹을 함께 키로 사용
            name_group_key = f"{person['이름']}_{person['그룹']}"
            if name_group_key not in seen:
                seen.add(name_group_key)
                unique_persons.append(person)
        
        # 추출된 인원수 확인
        extracted_count = len(unique_persons)
        
        # 추출된 인원수 출력
        st.write(f"명단에서 추출된 인원: {extracted_count}명")
        
        # 좌석 번호 생성
        low_seats = list(range(1, 20))
        high_seats = list(range(20, 222))
        chair_seats = [f"의자{i}" for i in range(1, 50)]  # 의자1-의자49

        # --- 특정 인원 좌석 범위 지정 ---
        special_seat_ranges = {
            "이인수": list(range(1, 71)),      # 1~70
            "이재길": list(range(1, 51)),      # 1~50
            "장한별": list(range(151, 222)),   # 150~221 (150번 이후)
        }
        special_seat_assignments = {}

        # 각 인원별로 좌석 미리 배정
        for name, seat_range in special_seat_ranges.items():
            person = next((p for p in unique_persons if p['이름'] == name), None)
            if person:
                available = [s for s in seat_range if s in low_seats or s in high_seats]
                if available:
                    chosen = random.choice(available)
                    special_seat_assignments[name] = chosen
                    # 좌석 리스트에서 제거
                    if chosen in low_seats:
                        low_seats.remove(chosen)
                    elif chosen in high_seats:
                        high_seats.remove(chosen)
                else:
                    st.error(f"{name}에게 배정할 수 있는 좌석이 없습니다!")
                    return None

        # 좌석 수와 명단 수 확인
        if len(unique_persons) > len(low_seats) + len(high_seats) + len(chair_seats):
            st.error(f"명단({len(unique_persons)}명)이 좌석 수({len(low_seats) + len(high_seats) + len(chair_seats)}개)보다 많습니다.")
            return None
            
        # 특정 그룹 분리 (7남, 8남, 15여, 16여)
        special_groups = ['7남', '8남', '15여', '16여']
        special_persons = [p for p in unique_persons if p['그룹'] in special_groups]
        regular_persons = [p for p in unique_persons if p['그룹'] not in special_groups]
        
        # 암호학적으로 안전한 난수 생성기를 사용하여 각 이름에 랜덤 값 할당
        for person in unique_persons:
            person['랜덤값'] = secrets.randbelow(1000000) / 1000000
        
        # 각 그룹별로 랜덤 값에 따라 정렬
        special_persons.sort(key=lambda x: x['랜덤값'])
        regular_persons.sort(key=lambda x: x['랜덤값'])
        
        # 일반 좌석 섞기
        random.shuffle(low_seats)
        random.shuffle(high_seats)
        
        # 특별 그룹에 높은 번호 좌석 배정, 일반 그룹에 나머지 좌석 배정
        needed_high_seats = min(len(special_persons), len(high_seats))
        
        # 높은 좌석이 부족한 경우, 일부 특별 그룹 사람은 일반 좌석 배정받음
        remaining_special = max(0, len(special_persons) - needed_high_seats)
        
        # 좌석 배정
        results = []
        
        # 특별 그룹에 높은 번호 좌석 배정
        for i in range(min(len(special_persons), needed_high_seats)):
            name = special_persons[i]['이름']
            if name in special_seat_assignments:
                results.append({
                    '이름': name,
                    '랜덤값': special_persons[i]['랜덤값'],
                    '당첨번호': special_seat_assignments[name]
                })
                continue
            results.append({
                '이름': name,
                '랜덤값': special_persons[i]['랜덤값'],
                '당첨번호': high_seats[i]
            })
        
        # 남은 높은 번호 좌석 + 낮은 번호 좌석을 일반 그룹 및 남은 특별 그룹에 배정
        remaining_high_seats = high_seats[needed_high_seats:]
        all_remaining_seats = remaining_high_seats + low_seats
        random.shuffle(all_remaining_seats)  # 다시 섞기
        
        # 남은 특별 그룹 사람들
        for i in range(remaining_special):
            idx = needed_high_seats + i
            name = special_persons[idx]['이름']
            if name in special_seat_assignments:
                results.append({
                    '이름': name,
                    '랜덤값': special_persons[idx]['랜덤값'],
                    '당첨번호': special_seat_assignments[name]
                })
                continue
            results.append({
                '이름': name,
                '랜덤값': special_persons[idx]['랜덤값'],
                '당첨번호': all_remaining_seats[i]
            })
        
        # 일반 그룹 사람들
        for i in range(len(regular_persons)):
            name = regular_persons[i]['이름']
            if name in special_seat_assignments:
                results.append({
                    '이름': name,
                    '랜덤값': regular_persons[i]['랜덤값'],
                    '당첨번호': special_seat_assignments[name]
                })
                continue
            if i + remaining_special < len(all_remaining_seats):
                results.append({
                    '이름': name,
                    '랜덤값': regular_persons[i]['랜덤값'],
                    '당첨번호': all_remaining_seats[i + remaining_special]
                })
            else:
                # 좌석이 부족하면 의자 배정
                chair_idx = i + remaining_special - len(all_remaining_seats)
                if chair_idx < len(chair_seats):
                    results.append({
                        '이름': name,
                        '랜덤값': regular_persons[i]['랜덤값'],
                        '당첨번호': chair_seats[chair_idx]
                    })
                else:
                    st.error(f"좌석 배정 중 오류가 발생했습니다: 남은 좌석이 없습니다.")
                    return None
        
        # 결과 데이터프레임 생성
        result_df = pd.DataFrame(results)
        
        # 이름 기준으로 정렬 (가나다순)
        result_df_sorted = result_df.sort_values(by='이름').reset_index(drop=True)
        
        # 필요한 의자 좌석 수 계산
        needed_chair_seats = sum(1 for item in results if isinstance(item['당첨번호'], str) and item['당첨번호'].startswith('의자'))
        needed_regular_seats = sum(1 for item in results if not (isinstance(item['당첨번호'], str) and item['당첨번호'].startswith('의자')))
        
        # 최종 인원수 검증
        if len(result_df) != extracted_count:
            st.warning(f"주의: 추출된 인원수({extracted_count})와 결과 인원수({len(result_df)})가 일치하지 않습니다!")
        
        return {
            'result_df': result_df_sorted,
            'names': [item['이름'] for item in unique_persons],
            'needed_regular_seats': needed_regular_seats,
            'needed_chair_seats': needed_chair_seats
        }

    except Exception as e:
        st.error(f"오류 발생: {e}")
        import traceback
        st.error(traceback.format_exc())
        return None



# 결과 엑셀 파일 생성 함수
def create_result_excel(results):
    # 결과 데이터프레임
    df = results['result_df']
    
    # 새 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "제비뽑기 결과"
    
    # 페이지 설정
    ws.page_setup.paperSize = 9  # A4 용지
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.horizontalCentered = True
    ws.print_options.horizontalCentered = True
    
    # 여백 설정
    ws.page_margins = PageMargins(bottom=0.4)
    
    # 맞춤 설정
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # 스타일 정의
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    light_blue_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    
    # 날짜 설정
    if 'file_date' in st.session_state:
        today = st.session_state.file_date.strftime('%Y년 %m월 %d일')
    else:
        today = datetime.now().strftime('%Y년 %m월 %d일')
    
    # 섹션별 행 수와 열 수
    rows_per_section = 30
    cols_per_section = 3
    
    total_persons = len(df)
    persons_per_section = rows_per_section * cols_per_section
    num_sections = (total_persons + persons_per_section - 1) // persons_per_section
    
    # 현재 행 위치
    current_row = 1
    
    # 섹션별로 데이터 추가
    for section_idx in range(num_sections):
        section_start_row = current_row
        
        # 제목 행
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        title_cell = ws.cell(row=current_row, column=1, value=f"제비뽑기 당첨 결과 {section_idx+1}")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 32
        current_row += 1
        
        # 날짜 행
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        date_cell = ws.cell(row=current_row, column=1, value=f"날짜: {today}")
        date_cell.font = Font(bold=True)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 24
        
        # (가나다순) 텍스트
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        sort_cell = ws.cell(row=current_row, column=5, value="(가나다순)")
        sort_cell.font = Font(bold=True)
        sort_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # 헤더 행
        headers = ["이 름", "당첨번호", "이 름", "당첨번호", "이 름", "당첨번호"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=current_row, column=i+1, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
        # 해당 섹션의 데이터 범위
        start_idx = section_idx * persons_per_section
        end_idx = min(start_idx + persons_per_section, total_persons)
        section_data = df.iloc[start_idx:end_idx].reset_index(drop=True)
        
        # 최대 행 인덱스 추적
        max_row_idx = -1
        
        # 섹션 데이터 추가
        for idx, row in section_data.iterrows():
            col_set = idx // rows_per_section
            row_idx = idx % rows_per_section
            max_row_idx = max(max_row_idx, row_idx)
            
            # 열 인덱스 계산
            col_idx = col_set * 2 + 1
            
            # 현재 데이터 행 위치
            data_row = current_row + row_idx
            
            # 이름 열과 당첨번호 열
            name_cell = ws.cell(row=data_row, column=col_idx)
            num_cell = ws.cell(row=data_row, column=col_idx + 1)
            
            # 스타일 설정
            name_cell.border = thin_border
            num_cell.border = thin_border
            num_cell.fill = light_blue_fill
            name_cell.font = Font(bold=True)
            num_cell.font = Font(bold=True)
            name_cell.alignment = Alignment(horizontal='center', vertical='center')
            num_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 데이터 설정
            name_cell.value = row['이름']
            num_cell.value = row['당첨번호']
            ws.row_dimensions[data_row].height = 22.80
        
        # 빈 데이터 처리
        if max_row_idx == -1:
            max_row_idx = 0
        
        # 섹션 마지막 행 계산
        section_end_row = current_row + max_row_idx
        
        # 섹션 테두리 추가
        for r in range(section_start_row, section_end_row + 1):
            for c in range(1, 7):
                if r == section_start_row or r == section_end_row or c == 1 or c == 6:
                    cell = ws.cell(row=r, column=c)
                    if cell.border:
                        # 테두리 처리 로직
                        if (r == section_start_row and c == 1):  # 좌상단 모서리
                            cell.border = Border(
                                left=Side(style='medium'),
                                right=cell.border.right,
                                top=Side(style='medium'),
                                bottom=cell.border.bottom
                            )
                        elif (r == section_start_row and c == 6):  # 우상단 모서리
                            cell.border = Border(
                                left=cell.border.left,
                                right=Side(style='medium'),
                                top=Side(style='medium'),
                                bottom=cell.border.bottom
                            )
                        elif (r == section_end_row and c == 1):  # 좌하단 모서리
                            cell.border = Border(
                                left=Side(style='medium'),
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=Side(style='medium')
                            )
                        elif (r == section_end_row and c == 6):  # 우하단 모서리
                            cell.border = Border(
                                left=cell.border.left,
                                right=Side(style='medium'),
                                top=cell.border.top,
                                bottom=Side(style='medium')
                            )
                        elif r == section_start_row:  # 상단 테두리
                            cell.border = Border(
                                left=cell.border.left,
                                right=cell.border.right,
                                top=Side(style='medium'),
                                bottom=cell.border.bottom
                            )
                        elif r == section_end_row:  # 하단 테두리
                            cell.border = Border(
                                left=cell.border.left,
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=Side(style='medium')
                            )
                        elif c == 1:  # 좌측 테두리
                            cell.border = Border(
                                left=Side(style='medium'),
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=cell.border.bottom
                            )
                        elif c == 6:  # 우측 테두리
                            cell.border = Border(
                                left=cell.border.left,
                                right=Side(style='medium'),
                                top=cell.border.top,
                                bottom=cell.border.bottom
                            )
                    else:
                        cell.border = medium_border
        
        # 다음 섹션 위치 업데이트
        current_row = section_end_row + 1
    
    # 열 너비 조정
    for i in range(1, 7):
        col_letter = get_column_letter(i)
        if i % 2 == 1:  # 홀수 열 (이름)
            ws.column_dimensions[col_letter].width = 15
        else:  # 짝수 열 (당첨번호)
            ws.column_dimensions[col_letter].width = 12
    
    # 당첨번호 순 결과 시트 추가
    ws_by_number = wb.create_sheet(title="당첨번호순 결과")
    
    # 헤더 설정
    header_cells = [
        ws_by_number.cell(row=1, column=1, value="당첨번호"),
        ws_by_number.cell(row=1, column=2, value="이름")
    ]
    
    for cell in header_cells:
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # 열 너비 설정
    ws_by_number.column_dimensions['A'].width = 12
    ws_by_number.column_dimensions['B'].width = 18
    
    # 데이터 정렬을 위한 함수
    def sort_key(item):
        number = item['당첨번호']
        # 숫자는 그대로 반환, 의자는 1000 이상의 숫자로 변환하여 정렬 순서 조정
        if isinstance(number, int) or str(number).isdigit():
            return int(number)
        elif isinstance(number, str) and number.startswith('의자'):
            try:
                # '의자1' -> 1001, '의자2' -> 1002 등으로 변환
                return 1000 + int(number.replace('의자', ''))
            except:
                return 9999  # 변환 실패 시 맨 뒤로
        else:
            return 9999  # 기타 형식은 맨 뒤로
    
    # 당첨번호 순으로 정렬
    result_by_number = sorted(df.to_dict('records'), key=sort_key)
    
    # 데이터 추가
    for idx, record in enumerate(result_by_number, 2):  # 2부터 시작 (헤더 다음 행)
        number_cell = ws_by_number.cell(row=idx, column=1, value=record['당첨번호'])
        name_cell = ws_by_number.cell(row=idx, column=2, value=record['이름'])
        
        # 스타일 설정
        number_cell.border = thin_border
        name_cell.border = thin_border
        number_cell.alignment = Alignment(horizontal='center', vertical='center')
        name_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 번호에 컬러 추가
        number_cell.fill = light_blue_fill
        
        # 행 높이 설정
        ws_by_number.row_dimensions[idx].height = 22.80
    
    # 좌석 배치표를 세 번째 시트로 추가
    try:
        # 내장된 좌석 배치표 열기
        src_wb = openpyxl.load_workbook(SEATING_CHART_PATH)
        src_ws = src_wb.active  # 첫 번째 시트 가져오기
        
        # 시트 복사 (서식 포함)
        ws2 = wb.create_sheet(title="좌석 배치표")
        
        # 페이지 설정 복사
        if src_ws.page_setup:
            ws2.page_setup.orientation = src_ws.page_setup.orientation
            ws2.page_setup.paperSize = src_ws.page_setup.paperSize
            ws2.page_setup.fitToHeight = src_ws.page_setup.fitToHeight
            ws2.page_setup.fitToWidth = src_ws.page_setup.fitToWidth
            ws2.page_setup.fitToPage = True  # 용지에 맞추기 설정 켜기
        
        # 페이지 여백 복사
        if src_ws.page_margins:
            ws2.page_margins = copy(src_ws.page_margins)
        
        # 인쇄 설정 복사
        ws2.print_options.horizontalCentered = src_ws.print_options.horizontalCentered
        ws2.print_options.verticalCentered = src_ws.print_options.verticalCentered
        
        # 원본 인쇄 영역이 있으면 복사, 없으면 전체 데이터 영역 사용
        if hasattr(src_ws, 'print_area') and src_ws.print_area:
            ws2.print_area = src_ws.print_area
        else:
            # 데이터가 있는 영역 계산
            min_row = 1
            min_col = 1
            max_row = max((c.row for c in src_ws._cells.keys()), default=1)
            max_col = max((c.column for c in src_ws._cells.keys()), default=1)
            
            # 인쇄 영역 설정
            min_col_letter = get_column_letter(min_col)
            max_col_letter = get_column_letter(max_col)
            ws2.print_area = f"{min_col_letter}{min_row}:{max_col_letter}{max_row}"

        # 셀 복사 (값과 서식 모두)
        for row_idx, row in enumerate(src_ws.rows, 1):
            for col_idx, cell in enumerate(row, 1):
                # 셀 값 복사
                new_cell = ws2.cell(row=row_idx, column=col_idx, value=cell.value)
                
                # 스타일 복사
                if cell.has_style:
                    # 폰트 복사
                    if cell.font:
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color
                        )
                    
                    # 테두리 복사
                    if cell.border:
                        new_cell.border = Border(
                            left=copy(cell.border.left) if cell.border.left else None,
                            right=copy(cell.border.right) if cell.border.right else None,
                            top=copy(cell.border.top) if cell.border.top else None,
                            bottom=copy(cell.border.bottom) if cell.border.bottom else None,
                            diagonal=copy(cell.border.diagonal) if cell.border.diagonal else None,
                            diagonalUp=cell.border.diagonalUp,
                            diagonalDown=cell.border.diagonalDown
                        )
                    
                    # 배경색 복사
                    if cell.fill and cell.fill.fill_type:
                        new_cell.fill = copy(cell.fill)
                    
                    # 정렬 복사
                    if cell.alignment:
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            textRotation=cell.alignment.textRotation,
                            wrapText=cell.alignment.wrapText,
                            shrinkToFit=cell.alignment.shrinkToFit,
                            indent=cell.alignment.indent
                        )
                    
                    # 숫자 형식 복사
                    new_cell.number_format = cell.number_format
                    
                    # 보호 설정 복사
                    if cell.protection:
                        new_cell.protection = Protection(
                            locked=cell.protection.locked,
                            hidden=cell.protection.hidden
                        )
        
        # 병합된 셀 복사
        for merged_range in src_ws.merged_cells.ranges:
            ws2.merge_cells(str(merged_range))
        
        # 행 높이 복사 - 원본 그대로
        for row_idx in range(1, src_ws.max_row + 1):
            if row_idx in src_ws.row_dimensions and src_ws.row_dimensions[row_idx].height:
                ws2.row_dimensions[row_idx].height = src_ws.row_dimensions[row_idx].height
        
        # 특정 열 너비 설정 (픽셀 기준)
        column_widths = {
            'A': pixels_to_excel_width(38),     # A열: 38픽셀
            'B': pixels_to_excel_width(61),     # B열: 61픽셀
            'C': pixels_to_excel_width(61),     # C열: 61픽셀
            'D': pixels_to_excel_width(15),     # D열: 15픽셀
            'E': pixels_to_excel_width(61),     # E열: 61픽셀
            'F': pixels_to_excel_width(61),     # F열: 61픽셀
            'G': pixels_to_excel_width(61),     # G열: 61픽셀
            'H': pixels_to_excel_width(15),     # H열: 15픽셀
            'I': pixels_to_excel_width(61),     # I열: 61픽셀
            'J': pixels_to_excel_width(61),     # J열: 61픽셀
            'K': pixels_to_excel_width(61),     # K열: 61픽셀
            'L': pixels_to_excel_width(15),     # L열: 15픽셀
            'M': pixels_to_excel_width(61),     # M열: 61픽셀
            'N': pixels_to_excel_width(61),     # N열: 61픽셀
            'O': pixels_to_excel_width(61),     # O열: 61픽셀
            'P': pixels_to_excel_width(15),     # P열: 15픽셀
            'Q': pixels_to_excel_width(61),     # Q열: 61픽셀
            'R': pixels_to_excel_width(61),     # R열: 61픽셀
            'S': pixels_to_excel_width(61)      # S열: 61픽셀
        }
        
        # 모든 열에 대해 너비 설정
        max_col = src_ws.max_column
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            
            if col_letter in column_widths:
                # 미리 계산된 특정 픽셀 값으로 설정
                ws2.column_dimensions[col_letter].width = column_widths[col_letter]
            elif col_letter in src_ws.column_dimensions and src_ws.column_dimensions[col_letter].width:
                # 다른 열은 원본과 동일한 비율로 설정
                # 원본 너비에 보정 계수 적용 (61/81 ≈ 0.75)
                ws2.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width * 0.75
            
            # 숨김 상태 복사
            if col_letter in src_ws.column_dimensions:
                ws2.column_dimensions[col_letter].hidden = src_ws.column_dimensions[col_letter].hidden
        
        # 이미지 복사 (있는 경우)
        if hasattr(src_ws, '_images'):
            for image in src_ws._images:
                try:
                    img_copy = Image(image.path)
                    img_copy.anchor = image.anchor
                    ws2.add_image(img_copy)
                except Exception as img_error:
                    print(f"이미지 복사 중 오류: {img_error}")
    
    except Exception as e:
        print(f"좌석 배치표 추가 중 오류 발생: {e}")
    
    # 엑셀 파일을 바이트로 변환
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# 페이지 설정
st.set_page_config(page_title="제비뽑기 프로그램", page_icon="🎯", layout="wide")

# CSS 스타일
st.markdown(""" 
<style>
/* 기본 모드 스타일 */
body { color: rgba(0, 0, 0, 0.9) !important; }
p, ol, ul, label, div { color: rgba(0, 0, 0, 0.9) !important; }
h1 { color: #000000 !important; text-align: center; margin-bottom: 2rem; }
h2, h3, h4 { color: #000000 !important; }

/* 다크 모드 스타일 */
@media (prefers-color-scheme: dark) {
    body { color: rgba(255, 255, 255, 0.9) !important; }
    p, ol, ul, label, div { color: rgba(255, 255, 255, 0.9) !important; }
    h1, h2, h3, h4 { color: #FFFFFF !important; }
}

.stButton > button {
    background-color: #4CAF50 !important;
    color: white !important;
    font-size: 20px !important;
    padding: 12px 24px !important;
    border-radius: 8px !important;
    border: none !important;
    cursor: pointer !important;
    margin: 20px 0 !important;
    display: block !important;
    width: 100% !important;
    transition: all 0.3s !important;
}
.stButton > button:hover {
    background-color: #3e8e41 !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important;
    transform: translateY(-2px) !important;
}

[data-testid="stDownloadButton"] > button {
    background-color: #4CAF50 !important;
    color: white !important;
    font-size: 20px !important;
    padding: 12px 24px !important;
    border-radius: 8px !important;
    border: none !important;
    cursor: pointer !important;
    margin: 20px 0 !important;
    display: block !important;
    width: 100% !important;
    transition: all 0.3s !important;
    height: auto !important;
    line-height: 1.6 !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background-color: #3e8e41 !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important;
    transform: translateY(-2px) !important;
}
[data-testid="stDownloadButton"] {
    margin-top: 20px !important;
    margin-bottom: 20px !important;
    display: block !important;
    width: 100% !important;
}

.css-1cpxqw2, [data-testid="stFileUploader"] {
    border: 2px dashed #4CAF50 !important;
    border-radius: 10px !important;
    padding: 30px !important;
    text-align: center !important;
    transition: all 0.3s !important;
    background-color: rgba(255, 255, 255, 0.05) !important;
    min-height: 280px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: center !important;
    align-items: center !important;
}
.css-1cpxqw2:hover, [data-testid="stFileUploader"]:hover {
    border-color: #ffffff !important;
    background-color: rgba(255, 255, 255, 0.1) !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.3) !important;
}

.download-container {
    padding: 20px !important;
    background-color: rgba(30, 30, 30, 0.6) !important;
    border-radius: 10px !important;
    margin: 0 !important;
    text-align: center !important;
    border: 1px solid rgba(255, 255, 255, 0.2) !important;
    min-height: 280px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: center !important;
    align-items: center !important;
    width: 100% !important;
}
.equal-height-container {
    height: 100% !important;
    display: flex !important;
    flex-direction: column !important;
}

.info-box {
    background-color: rgba(30, 30, 30, 0.6) !important;
    padding: 15px !important;
    border-radius: 8px !important;
    margin: 20px 0 !important;
    border-left: 5px solid #4CAF50 !important;
    color: rgba(0, 0, 0, 0.9) !important;  /* 기본 모드에서 텍스트 색상 변경 */
}
.element-container.st-success {
    background-color: rgba(76, 175, 80, 0.1) !important;
    padding: 15px !important;
    border-radius: 8px !important;
    border-left: 5px solid #4CAF50 !important;
    margin: 20px 0 !important;
}
.result-summary {
    margin-top: 20px !important;
    background-color: rgba(30, 30, 30, 0.5) !important;
    padding: 15px !important;
    border-radius: 8px !important;
    border-left: 5px solid #4CAF50 !important;
}
</style>
""", unsafe_allow_html=True)

# 제목과 설명
st.title("🎯 신학교 제비뽑기 프로그램")
st.markdown("""
<div class="info-box">
<h3>사용 방법</h3>
<ol>
<li>명단이 있는 엑셀 파일(.xlsx, .xls)을 업로드하세요.</li>
<li>'제비뽑기 실행' 버튼을 클릭하세요.</li>
<li>결과가 생성되면 '결과 파일 다운로드' 버튼을 클릭하여 저장하세요.</li>
<li>Excel 파일이 한 페이지에 모든 결과가 나타나도록 설정되었습니다.</li>
</ol>
</div>
""", unsafe_allow_html=True)

# 화면 2단 분할
col1, col2 = st.columns([1, 1])

with col1:
    st.markdown('<div class="equal-height-container">', unsafe_allow_html=True)
    
    # 파일 업로드 위젯
    uploaded_file = st.file_uploader("명단이 있는 엑셀 파일을 업로드하세요", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        st.success(f"파일 '{uploaded_file.name}'이 업로드되었습니다.")
        
        # 파일 이름에서 날짜 추출
        file_date = None
        filename_without_ext = os.path.splitext(uploaded_file.name)[0]

        # 날짜 형식 확인 및 추출 시도
        date_formats = [
            '%Y-%m-%d',  # 2025-04-01
            '%Y%m%d',    # 20250401
            '%Y_%m_%d',  # 2025_04_01
            '%y%m%d',    # 250401 <- 이 형식 추가
            '%m%d',      # 0401 (당해 연도 사용)
            '%m-%d',     # 04-01
            '%m_%d'      # 04_01
        ]

        for date_format in date_formats:
            try:
                if date_format == '%y%m%d':  # YY년MM월DD일 형식
                    extracted_date = datetime.strptime(filename_without_ext, date_format)
                    # 20XX년으로 설정
                    if extracted_date.year < 100:
                        extracted_date = extracted_date.replace(year=extracted_date.year + 2000)
                    file_date = extracted_date
                    break
                elif len(date_format) == 5:  # %m%d 형식인 경우 
                    extracted_date = datetime.strptime(filename_without_ext, date_format)
                    # 현재 연도 추가
                    current_year = datetime.now().year
                    extracted_date = extracted_date.replace(year=current_year)
                    file_date = extracted_date
                    break
                else:
                    extracted_date = datetime.strptime(filename_without_ext, date_format)
                    file_date = extracted_date
                    break
            except ValueError:
                pass

        # 직접 패턴 매칭 시도 (위 형식이 모두 실패한 경우)
        if file_date is None:
            # 예: 250409 형식 처리
            if len(filename_without_ext) == 6 and filename_without_ext.isdigit():
                try:
                    yy = int(filename_without_ext[0:2])
                    mm = int(filename_without_ext[2:4])
                    dd = int(filename_without_ext[4:6])
                    
                    if 1 <= mm <= 12 and 1 <= dd <= 31:  # 날짜 유효성 검사
                        year = 2000 + yy  # 20XX년으로 변환
                        file_date = datetime(year, mm, dd)
                        st.info(f"파일명에서 날짜를 추출했습니다: {file_date.strftime('%Y년 %m월 %d일')}")
                except:
                    pass

        # 날짜 추출 실패 시 현재 날짜 사용
        if file_date is None:
            file_date = datetime.now()
            st.warning("파일명에서 날짜를 추출할 수 없어 현재 날짜를 사용합니다.")

        # 추출된 날짜를 세션 상태에 저장
        st.session_state.file_date = file_date

        
        # 제비뽑기 실행 버튼
        if st.button("제비뽑기 실행"):
            with st.spinner("제비뽑기 진행 중..."):
                results = create_random_seating_assignment(uploaded_file)
                
                if results:
                    st.session_state.results = results
                    st.session_state.excel_data = create_result_excel(results)
                    st.session_state.execution_completed = True
                    
                    # 결과 요약
                    needed_regular = results['needed_regular_seats']
                    needed_chair = results['needed_chair_seats']
                    total_people = needed_regular + needed_chair
                    
                    st.success(f"✅ 제비뽑기 완료! 총 {total_people}명 배정 ({needed_regular}개 일반 좌석, {needed_chair}개 의자 좌석)")
    
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="equal-height-container" style="width:100%;">', unsafe_allow_html=True)
    
    if 'execution_completed' in st.session_state and st.session_state.execution_completed:
        st.markdown(""" 
        <div class="download-container">
        <h3>결과 다운로드</h3>
        <p>제비뽑기가 완료되었습니다!</p>
        """, unsafe_allow_html=True)
        
        # 날짜 형식의 파일명 생성
        if 'file_date' in st.session_state:
            date_str = st.session_state.file_date.strftime('%Y%m%d')
            file_name = f"제비뽑기_결과_{date_str}.xlsx"
        else:
            # 폴백: 현재 날짜 사용
            file_name = f"제비뽑기_결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Excel 파일 다운로드 버튼
        st.download_button(
            label="📥 결과 파일 다운로드",
            data=st.session_state.excel_data,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel",
            help="결과 Excel 파일을 다운로드합니다.",
            use_container_width=True
        )
        
        st.write(f"일반 좌석: {st.session_state.results['needed_regular_seats']}개")
        st.write(f"의자 좌석: {st.session_state.results['needed_chair_seats']}개")
        
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown(""" 
        <div class="download-container">
        <h3>결과 다운로드</h3>
        <p>왼쪽에서 파일을 업로드하고 제비뽑기를 실행하면 여기에 다운로드 버튼이 나타납니다.</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)

# 프로그램 설명
with st.expander("제비뽑기 프로그램 상세 설명"):
    st.markdown("""
    ### 제비뽑기 프로그램 특징
    
    **좌석 배정 방식:**
    - 일반 좌석(1~221)이 랜덤하게 배정됩니다.
    - 인원이 221명을 초과하는 경우에만 의자 좌석이 배정됩니다.
    - 의자 좌석은 의자1부터 순차적으로 필요한 만큼만 배정됩니다.
    
    **결과 파일 형식:**
    - Excel 파일로 다운로드됩니다.
    - 3개의 시트가 있습니다: 제비뽑기 결과(가나다순), 당첨번호순 결과, 좌석 배치표
    - 각 섹션별로 제목과 헤더가 추가되어 구분이 용이합니다.
    - 세로 방향 인쇄로 설정되어 있으며, 페이지 여백이 가로 가운데 맞춤으로 조정되었습니다.
    - 모든 텍스트는 굵게 처리되고 중앙 정렬됩니다.
    - 당첨번호 열은 연한 파란색 배경으로 표시됩니다.
    """)

# 푸터
st.markdown("---")
st.markdown("<p style='text-align: center; color: rgba(250, 250, 250, 0.7);'>© 2025 신학교 제비뽑기 프로그램 | 제작자: 여치형</p>", unsafe_allow_html=True)
